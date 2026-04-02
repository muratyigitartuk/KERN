from __future__ import annotations

import re
from pathlib import Path
from typing import Any


class SpreadsheetParser:
    """Parse CSV and Excel spreadsheets into structured data for RAG indexing and querying."""

    SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

    @staticmethod
    def parse_csv(path: Path | str) -> list[dict[str, Any]]:
        import csv

        path = Path(path)
        rows: list[dict[str, Any]] = []
        with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(dict(row))
        return rows

    @staticmethod
    def parse_excel(path: Path | str) -> dict[str, list[dict[str, Any]]]:
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel ingestion. Install it with: pip install openpyxl") from exc

        path = Path(path)
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                sheets[sheet_name] = []
                continue
            headers = [str(cell) if cell is not None else f"col_{i}" for i, cell in enumerate(rows[0])]
            data_rows: list[dict[str, Any]] = []
            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                data_rows.append(
                    {
                        headers[i]: (cell if cell is not None else "")
                        for i, cell in enumerate(row)
                        if i < len(headers)
                    }
                )
            sheets[sheet_name] = data_rows
        return sheets

    @staticmethod
    def to_natural_language(data: list[dict[str, Any]], title: str = "Spreadsheet", max_rows: int = 200) -> str:
        if not data:
            return f"{title}: No data."
        headers = list(data[0].keys())
        lines = [f"# {title}", f"Columns: {', '.join(headers)}", f"Rows: {len(data)}", ""]
        for i, row in enumerate(data[:max_rows]):
            parts = [f"{k}: {v}" for k, v in row.items() if v not in (None, "")]
            if parts:
                lines.append(f"Row {i + 1}: {', '.join(parts)}")
        if len(data) > max_rows:
            lines.append(f"... and {len(data) - max_rows} more rows.")
        return "\n".join(lines)

    @staticmethod
    def query_dataframe(data: list[dict[str, Any]], query: str) -> str:
        if not data:
            return "No data available."
        query_lower = query.lower()
        headers = list(data[0].keys())
        filtered = SpreadsheetParser._apply_filters(data, query_lower)
        group_by = SpreadsheetParser._detect_group_by(headers, query_lower)
        target_column = SpreadsheetParser._match_column(headers, query_lower)
        if "count" in query_lower or "how many" in query_lower or "anzahl" in query_lower or "wie viele" in query_lower:
            if group_by:
                grouped = SpreadsheetParser._group_counts(filtered, group_by)
                return f"Counts by {group_by}: " + ", ".join(f"{key}: {value}" for key, value in grouped[:6])
            if filtered is not data:
                return f"Matched rows: {len(filtered)}"
            return f"Total rows: {len(data)}"
        if group_by and any(token in query_lower for token in {"sum", "total", "average", "avg", "mean", "summe", "gesamt", "durchschnitt", "mittelwert"}) and target_column:
            grouped_values = SpreadsheetParser._group_numeric(filtered, group_by, target_column)
            if grouped_values:
                if any(token in query_lower for token in {"average", "avg", "mean", "durchschnitt", "mittelwert"}):
                    return f"Averages by {group_by}: " + ", ".join(
                        f"{key}: {sum(values) / len(values):,.2f}" for key, values in grouped_values[:6] if values
                    )
                return f"Sums by {group_by}: " + ", ".join(
                    f"{key}: {sum(values):,.2f}" for key, values in grouped_values[:6] if values
                )
        if any(token in query_lower for token in {"sum", "total", "summe", "gesamt"}):
            totals = SpreadsheetParser._numeric_aggregates(filtered, headers, "sum", target_column)
            if totals:
                return "Sums: " + ", ".join(f"{col}: {value:,.2f}" for col, value in totals.items())
        if any(token in query_lower for token in {"average", "avg", "mean", "durchschnitt", "mittelwert"}):
            avgs = SpreadsheetParser._numeric_aggregates(filtered, headers, "avg", target_column)
            if avgs:
                return "Averages: " + ", ".join(f"{col}: {value:,.2f}" for col, value in avgs.items())
        if "show" in query_lower or "list" in query_lower or "where" in query_lower or "zeige" in query_lower or "zeig" in query_lower or "liste" in query_lower or "wo" in query_lower:
            preview = SpreadsheetParser._preview_rows(filtered, headers)
            if filtered is not data:
                return f"Filtered rows: {len(filtered)}\n{preview}"
            return preview
        return SpreadsheetParser._preview_rows(filtered, headers)

    @staticmethod
    def extract_text_from_csv(path: Path | str) -> str:
        rows = SpreadsheetParser.parse_csv(path)
        return SpreadsheetParser.to_natural_language(rows, title=Path(path).stem)

    @staticmethod
    def extract_text_from_excel(path: Path | str) -> str:
        sheets = SpreadsheetParser.parse_excel(path)
        parts: list[str] = []
        for sheet_name, rows in sheets.items():
            parts.append(SpreadsheetParser.to_natural_language(rows, title=sheet_name))
        return "\n\n".join(parts) if parts else "Empty workbook."

    @staticmethod
    def _preview_rows(data: list[dict[str, Any]], headers: list[str]) -> str:
        lines = [f"Data has {len(data)} rows and columns: {', '.join(headers)}."]
        for row in data[:5]:
            parts = [f"{k}={v}" for k, v in row.items() if v not in (None, "")]
            lines.append("  " + ", ".join(parts[:6]))
        if len(data) > 5:
            lines.append(f"  ... and {len(data) - 5} more rows.")
        return "\n".join(lines)

    @staticmethod
    def _match_column(headers: list[str], query_lower: str) -> str | None:
        normalized = {header.lower(): header for header in headers}
        for prefix in ("sum of ", "total of ", "average of ", "avg of ", "mean of ", "summe von ", "gesamt ", "durchschnitt von ", "mittelwert von ", "anzahl von "):
            if prefix in query_lower:
                token = query_lower.split(prefix, 1)[1].split(" where ", 1)[0].split(" wo ", 1)[0].split(" gruppiert ", 1)[0].strip()
                for lowered, original in normalized.items():
                    if token == lowered or token in lowered:
                        return original
        for lowered, original in normalized.items():
            if lowered in query_lower:
                return original
        return None

    @staticmethod
    def _apply_filters(data: list[dict[str, Any]], query_lower: str) -> list[dict[str, Any]]:
        if " where " not in query_lower and " wo " not in query_lower:
            return data
        if " where " in query_lower:
            clause = query_lower.split(" where ", 1)[1]
        else:
            clause = query_lower.split(" wo ", 1)[1]
        for operator in (" contains ", " enthält ", "=", " is "):
            if operator in clause:
                left, right = clause.split(operator, 1)
                column_hint = left.strip()
                raw_value = right.split(" and ", 1)[0].strip().strip("'\"")
                contains_mode = operator.strip() in ("contains", "enthält")
                break
        else:
            return data
        if not column_hint or not raw_value:
            return data
        headers = list(data[0].keys())
        column = SpreadsheetParser._match_column(headers, column_hint)
        if column is None:
            return data
        target = raw_value.lower()
        if contains_mode:
            return [row for row in data if target in str(row.get(column, "")).strip().lower()]
        return [row for row in data if str(row.get(column, "")).strip().lower() == target]

    @staticmethod
    def _detect_group_by(headers: list[str], query_lower: str) -> str | None:
        match = re.search(r"group by ([a-z0-9_ ]+)", query_lower)
        if not match:
            match = re.search(r"gruppiert nach ([a-z0-9_ äöü]+)", query_lower)
        if not match:
            match = re.search(r"pro ([a-z0-9_ äöü]+)", query_lower)
        if not match:
            return None
        column_hint = match.group(1).strip().split(" where ", 1)[0].split(" wo ", 1)[0].strip()
        return SpreadsheetParser._match_column(headers, column_hint)

    @staticmethod
    def _group_counts(data: list[dict[str, Any]], column: str) -> list[tuple[str, int]]:
        counts: dict[str, int] = {}
        for row in data:
            key = str(row.get(column, "") or "unknown")
            counts[key] = counts.get(key, 0) + 1
        return sorted(counts.items(), key=lambda item: (-item[1], item[0]))

    @staticmethod
    def _group_numeric(data: list[dict[str, Any]], group_by: str, target_column: str) -> list[tuple[str, list[float]]]:
        grouped: dict[str, list[float]] = {}
        for row in data:
            value = SpreadsheetParser._coerce_numeric(row.get(target_column, ""))
            if value is None:
                continue
            key = str(row.get(group_by, "") or "unknown")
            grouped.setdefault(key, []).append(value)
        return sorted(grouped.items(), key=lambda item: item[0])

    @staticmethod
    def _numeric_aggregates(data: list[dict[str, Any]], headers: list[str], mode: str, target_column: str | None) -> dict[str, float]:
        columns = [target_column] if target_column else headers
        results: dict[str, float] = {}
        for col in columns:
            values: list[float] = []
            for row in data:
                value = SpreadsheetParser._coerce_numeric(row.get(col, ""))
                if value is not None:
                    values.append(value)
            if not values:
                continue
            if mode == "sum":
                results[col] = sum(values)
            else:
                results[col] = sum(values) / len(values)
        return results

    @staticmethod
    def _coerce_numeric(value: Any) -> float | None:
        if value in (None, ""):
            return None
        text = str(value).strip()
        if not text:
            return None
        text = text.replace("€", "").replace("$", "").replace("£", "").replace("EUR", "").replace("USD", "").replace("GBP", "")
        text = text.replace(" ", "")
        if text.count(",") == 1 and text.count(".") == 0:
            text = text.replace(",", ".")
        elif text.count(",") >= 1 and text.count(".") >= 1:
            text = text.replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
